#!/usr/bin/env python3

# USAGE:   python calcSequinAddition.py


from pathlib import Path
import pandas as pd
import numpy as np
import sys
import os
import openpyxl
import shutil
# from openpyxl import load_workbook
import xlwings as xw
from datetime import datetime
import matplotlib.pyplot as plt
from PyPDF2 import PdfFileMerger

# call exists() function named 'file_exists'
from os.path import exists as file_exists


# define list of destination well positions for a 96-well and 384-well plates
well_list_96w = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'A3', 'B3', 'C3',
                 'D3', 'E3', 'F3', 'G3', 'H3', 'A4', 'B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'H4', 'A5', 'B5', 'C5', 'D5', 'E5', 'F5', 'G5',
                 'H5', 'A6', 'B6', 'C6', 'D6', 'E6', 'F6', 'G6', 'H6', 'A7', 'B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7', 'A8', 'B8', 'C8',
                 'D8', 'E8', 'F8', 'G8', 'H8', 'A9', 'B9', 'C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'A10', 'B10', 'C10', 'D10', 'E10', 'F10', 'G10',
                 'H10', 'A11', 'B11', 'C11', 'D11', 'E11', 'F11', 'G11', 'H11', 'A12', 'B12', 'C12', 'D12', 'E12', 'F12', 'G12', 'H12']


##########################
##########################
def getMatchedFiles(dirname, prefix):
    # create empty list to hold names of files where DNA conc with prefix has matching
    # density and volume check files
    list_matched_files = []

    # loop through all files in directory and find DNA conc files with prefix
    for file in os.listdir(dirname):
        if file.startswith(prefix):
            
            # confirm the DNA conc file has matching prefix and suffix
            # e.g.  preSIP0999pre.txt  instead of preSIP0999post.txt
            check_dna_conc_format = file.replace(".txt","")
            
            if not check_dna_conc_format.endswith(prefix):
                print(f'\n Problem with DNA conc format for file {file}. Aborting process\n\n')
                sys.exit()
            

            # extract plate id from file name from density file name and add to plate id list
            plateid = file.replace(prefix, "")
            plateid = plateid.replace(".txt", "")

            # create volume check file name
            vol_file = dirname+"/"+plateid+".CSV"

            # create dna conc file name
            dens_file = dirname+"/"+plateid+".xlsx"

            # check if DNA conc file with prefix has corresponding density (.xlsx) and volume (.csv) files
            if (file_exists(dens_file)):
                if (file_exists(vol_file)):
                    list_matched_files.append(file)
                else:
                    print(
                        f'\n\n Warning: File was not found\nVolume file{vol_file}\n\n')
                    continue
            else:
                print(
                    f'\n\n Warning: File was not found\nDensity file {dens_file} \n\n')
                continue

        else:
            continue

    # quit script if directory doesn't contain any matched sets of density, conc, volume files
    if len(list_matched_files) == 0:
        print(
            f"\n\n Did not find any matche sets of desnity and volume info corresponding with DNA conc files that have {prefix} in name.  Aborting program\n\n")
        sys.exit()

    return list_matched_files
##########################
##########################

##########################
##########################


def getVolumes(my_dirname, my_plateid):
    vol_path = my_dirname+"/" + my_plateid+".CSV"

    # import corrsponding volume check file
    my_volume_df = pd.read_csv(vol_path, header=0, usecols=['RACKID', 'TUBE', 'VOLAVG'], converters={
        'TUBE': str, 'RACKID': str}, skip_blank_lines=True)
    # volume_df = pd.read_csv(plateid+".CSV", header=0, usecols=['RACKID', 'TUBE', 'VOLAVG'], converters={
    #     'TUBE': str, 'RACKID': str})

    # replace missing values with 0 ONLY for well A01.  This seems to be a glitch
    # in instrument used to measure well volumes
    my_volume_df['VOLAVG'] = np.where(
        ((my_volume_df['TUBE'] == 'A01') & (my_volume_df['VOLAVG'].isnull())), 0, my_volume_df['VOLAVG'])

    if my_volume_df['VOLAVG'].isnull().values.any():
        print(
            f"\n\nERROR\n\nThe volume plate for {my_plateid} is missing data.  Aborting script\n\n")
        sys.exit()

    my_volume_df['VOLAVG'] = my_volume_df['VOLAVG'].astype(int)

    # loop to reformat well positions, e.g. A01 --> A1, B01 --> B1
    new_pos_list = []
    for index, row in my_volume_df.iterrows():
        # print(row['TUBE'])
        well = row['TUBE']
        if well[1] == '0':
            well = well.replace('0', '')

        new_pos_list.append(well)

    # replace old well format with new format
    my_volume_df["TUBE"] = new_pos_list

    return my_volume_df

##########################
##########################

##########################
##########################


def getDensity(my_dirname, my_plateid):
    # Read in density file
    dens_path = my_dirname+"/"+my_plateid+".xlsx"

    # # solves a problem that arose where formulas in density.xlsx files would no longer import values
    # # for some reason, problem is solved by opening, saving, and closing the files before reading into pandas
    # # the 4 lines below automaticaly open, save, and close the density files
    # app = xl.App(visible=False)
    # book = app.books.open(dens_path)
    # book.save()
    # app.kill()

    my_density_df = pd.read_excel(dens_path, header=0, engine=("openpyxl"), usecols=['Plate barcode', 'Sample barcode', 'Well Pos', 'Fraction #', 'Density', 'Spike-in Mass (pg)'],
                                  converters={'Plate barcode': str, 'Sample barcode': str, 'Fraction #': int})

    # check if desnity .xlsx file already hase values ented in Spike-in Mass (pg) column
    if not(my_density_df['Spike-in Mass (pg)'].isnull().all()):

        keep_going = str(
            input(f'Density file {my_plateid}.xlsx already has values in column Spike-in Mass (pg)\n\nDo you want to overwrite these data?  (y/n)  ') or 'n')

        if (keep_going == 'Y' or keep_going == 'y'):
            print("Ok, we'll keep going and overwrite existing sequin mass data\n\n")

        elif (keep_going == 'N' or keep_going == 'n'):
            print(
                'Ok, aborting script.\n\n')
            sys.exit()
        else:
            print("Sorry, you must choose 'Y' or 'N' next time. \n\nAborting\n\n")
            sys.exit()

    # get rid of spike-in mass column
    my_density_df.drop(columns=['Spike-in Mass (pg)'], inplace=True)

    # make a copy of df that include the empty row between the two samples in density plates
    my_density_emptyrow_df = my_density_df.copy()

    my_density_df = my_density_df[my_density_df['Density'].notna()]

    my_density_df['Density'] = my_density_df['Density'].astype(float)

    my_density_df = my_density_df.round({'Density': 4})

    return my_density_df, my_density_emptyrow_df

##########################
##########################

##########################
##########################


def getConc(my_dirname, my_dna):

    conc_path = my_dirname+"/" + my_dna

    tmp_conc_file = 'tmp_conc.txt'

    with open(tmp_conc_file, 'w') as outfile:
        with open(conc_path, 'r') as file:
            for line in file:
                if not line.isspace():
                    outfile.write(line)

    my_DNA_conc_df = pd.read_csv(tmp_conc_file, sep='\t', header=1, usecols=['Well ID', 'Well', '[Concentration]'],
                                 converters={'[Concentration]': str}, skip_blank_lines=True)

    # remove empty lines with NaN
    my_DNA_conc_df.dropna(inplace=True)

    # remove blanks and standards from dataframe and keep only rows with "SPL" in well ID
    my_DNA_conc_df = my_DNA_conc_df.loc[my_DNA_conc_df['Well ID'].str.contains(
        'SPL')]

    my_DNA_conc_df.drop('Well ID', inplace=True, axis=1)

    # remove "<" character from dna conc for samples below detection limit
    my_DNA_conc_df["[Concentration]"] = my_DNA_conc_df["[Concentration]"].str.replace(
        '<', '')

    # replace dna conc values <= 0.001 with 0.001 ng/ul.  Clarity/ITS won't accept a conc <= 0
    my_DNA_conc_df["[Concentration]"] = np.where(my_DNA_conc_df["[Concentration]"].astype(
        float) <= 0.001, 0.001, my_DNA_conc_df["[Concentration]"].astype(float))

    # round DNA concentration
    my_DNA_conc_df = my_DNA_conc_df.round({'[Concentration]': 3})

    # append my_DNA_conc_df to larger DF of all files

    try:
        os.remove(tmp_conc_file)

    except:
        print(f'\nError deleting {tmp_conc_file}')

    return my_DNA_conc_df

##########################
##########################


##########################
##########################
def getMergedPlates(DNA_conc_df, volume_df, density_df, plateid, list_merged_plates):
    # create dictionarly to hold merged df for each plateid
    # the plate id is the key, and the merged df is the value
    d = {}

    # merge denisty and volume dataframes
    d[plateid] = density_df.merge(volume_df, how='outer', left_on=['Well Pos', 'Plate barcode'],
                                  right_on=['TUBE', 'RACKID'])

    # generate error if there is a mismatch in wells or plate id between density and volume files
    if (d[plateid]["TUBE"].isnull().values.any()):
        print("\n\n")
        print(d[plateid])
        print(
            '\n\n Wells and plated id do not match. Aborting. Check error file. \n\n')
        d[plateid].to_csv('error.desnity.volume.merge.csv', index=False)

        sys.exit()

    # merge dnc conc dataframe with density and volume df
    d[plateid] = d[plateid].merge(DNA_conc_df, how='inner', left_on='Well Pos',
                                  right_on='Well')

    # generate error if there is a mismatch in wells or plate id between dna conc file and merged density+volume df
    if (d[plateid]["TUBE"].isnull().values.any()):
        print("\n\n")
        print(d[plateid])
        print(
            '\n\n Wells and plated id do not match. Aborting. Check error file. \n\n')
        d[plateid].to_csv('error.desnity.volume.merge.csv', index=False)

        sys.exit()

    # create list of successfully merged plate IDs
    list_merged_plates.append(plateid)

    # # quit script if were not able to successfully merge any density, conc, volume files
    # if len(list_merged_plates) == 0:
    #     print("\n\n Did not sucessfully merge sets of files\n\n")
    #     sys.exit()

    # else:  # sort list of sucessfully merged fraction plates
    #     list_merged_plates.sort()
    #     print("\nList of merged sample plates:\n\n", list_merged_plates)

    return d[plateid], list_merged_plates
##########################
##########################


##########################
##########################
def calcSequinMassRange(hi_sequin_working_conc, lo_sequin_working_conc, min_trans_vol, max_trans_vol):

    # divide by 1000 to convert from pg/ul to ng/ul

    hi_range = [(hi_sequin_working_conc*min_trans_vol/1000),
                (hi_sequin_working_conc*max_trans_vol/1000)]

    lo_range = [(lo_sequin_working_conc*min_trans_vol/1000),
                (lo_sequin_working_conc*max_trans_vol/1000)]

    return hi_range, lo_range
##########################
##########################


##########################
##########################
def enterFractionMetadata():
    # ask user for total number of fractions collected
    total_fractions = float(
        input("Enter the total # fractions collected (default 24): ") or 24)

    if (total_fractions <= 0):
        print('\n\nError.  Fractions must be >0.  Aborting.\n\n')
        sys.exit()

    # ask user for target % for sequins
    percent_sequin = float(
        input("Sequin mass should be what % of total mass? (default 1): ") or 1)

    if ((percent_sequin < 0) | (percent_sequin >= 100)):
        print('\n\nError.  The % sequin should be  >=0 and <100.  Aborting.\n\n')
        sys.exit()

    # ask user for volume of cscl fraction used to determine DNA conc
    cscl_vol = float(
        input("Enter the uL's of fraction used to measure DNA conc  (default 2.4): ") or 2.4)

    if ((cscl_vol <= 0) | (percent_sequin > 10)):
        print('\n\nError.  Fraction volume should be >0 and <=10uL  Aborting.\n\n')
        sys.exit()

    # ask user for conc of sequin working stock
    hi_sequin_working_conc = float(
        input("Enter concentration of HIGH sequin working stock (default 70 pg/uL): ") or 70)

    # ask user for conc of sequin working stock
    lo_sequin_working_conc = float(
        input("Enter concentration of LOW sequin working stock (default 15 pg/uL): ") or 15)

    if ((lo_sequin_working_conc <= 0) | (hi_sequin_working_conc <= 0)):
        print('\n\nError.  Sequin_working_conc >0.  Aborting.\n\n')
        sys.exit()

    # ask user for minimum transfer volume
    min_trans_vol = float(
        input("Enter minium transfer volume for sequin addition (default 2.4 uL): ") or 2.4)

    if (min_trans_vol <= 0):
        print('\n\nError.  minimum transfer volume must be >0.  Aborting.\n\n')

    # ask user for minimum transfer volume
    max_trans_vol = float(
        input("Enter maximum transfer volume for sequin addition (default 35 uL): ") or 35)

    return total_fractions, percent_sequin, cscl_vol, hi_sequin_working_conc, lo_sequin_working_conc, min_trans_vol, max_trans_vol
##########################
##########################


##########################
##########################
def calcCorrectionFactor(cscl_vol):
    # calculate degree of DNA conc underestimation due
    # to interfernece by cscl in quant-it kit
    # number will be negative
    inhibit = (cscl_vol * -0.041)+0.028

    # convert inhibit to postive and add 1 so that
    # mulitplying measured DNA conc by the correction factor
    # compensates for the inhibition
    correction_fractor = 1+abs(inhibit)

    return correction_fractor

##########################
##########################



##########################
##########################


def addSequinVol(my_merged_df, total_fractions, percent_sequin, cscl_vol, hi_sequin_working_conc, lo_sequin_working_conc, min_trans_vol, max_trans_vol):

    # get correction factor for CsCl inhibition of quant-it kit
    # for  measuring DNA conc
    correction_factor = calcCorrectionFactor(cscl_vol)

    my_merged_df['sample_mass_(ng)'] = my_merged_df['VOLAVG'] * \
        my_merged_df['[Concentration]'] * correction_factor

    my_merged_df['sequin_mass_(ng)'] = my_merged_df['sample_mass_(ng)'] * \
        (percent_sequin/100)

    # set minimum sequin mass at 0.1ng aka 100pg
    my_merged_df['sequin_mass_(ng)'] = np.where(
        my_merged_df['sequin_mass_(ng)'] < 0.1, 0.1, my_merged_df['sequin_mass_(ng)'])

    # calculate the range of sequin masses that can be added from hi and lo sequin working stocks
    hi_sequin_range, lo_sequin_range = calcSequinMassRange(
        hi_sequin_working_conc, lo_sequin_working_conc, min_trans_vol, max_trans_vol)

    # determine if should use hi or low concentration of sequin working stock
    my_merged_df['sequin_stock'] = np.where(
        my_merged_df['sequin_mass_(ng)'] >= hi_sequin_range[0], 'HI', 'lo')

    # add concentration (pg/uL) of hi and lo sequin working stocks
    my_merged_df['sequin_working_stock_conc_(pg/uL)'] = np.where(
        my_merged_df['sequin_stock'] == "HI", hi_sequin_working_conc, lo_sequin_working_conc)

    # calculated volum of sequin working stock needed
    # divide stock conc by 1000 to convert pg/uL to ng/uL
    my_merged_df['sequin_vol_(uL)'] = my_merged_df['sequin_mass_(ng)'] / \
        (my_merged_df['sequin_working_stock_conc_(pg/uL)']/1000)

    # adjust sequin transfer vol if it exceeds the max transfer volume
    my_merged_df['sequin_vol_(uL)'] = np.where(
        my_merged_df['sequin_vol_(uL)'] > max_trans_vol, max_trans_vol, my_merged_df['sequin_vol_(uL)'])

    # adjust sequin transfer vol if it falls below minimum transfer volume
    my_merged_df['sequin_vol_(uL)'] = np.where(
        my_merged_df['sequin_vol_(uL)'] < min_trans_vol, min_trans_vol, my_merged_df['sequin_vol_(uL)'])

    # calculate the actual mass of sequins transferred for fractions where target mass exceeded max transfer volume
    my_merged_df['actual_sequin_mass_(ng)'] = (my_merged_df['sequin_vol_(uL)'] *
                                               my_merged_df['sequin_working_stock_conc_(pg/uL)']).astype(int)

    my_merged_df = my_merged_df.round({'sequin_vol_(uL)': 1})
    
    sample_list = my_merged_df['Sample barcode'].unique().tolist()

    # select only wells where fractions were collected
    # row index 40 = well position A7... this is result
    # of removing standard curve samples from dna_conc_df
    # if else statements used in case a density .xlsx sheet has only
    # one sample per plate instead of two

    # if (my_merged_df.shape[0] <= 40) and (len(my_merged_df['Sample barcode'].unique().tolist()) == 1):
    #     my_merged_df = my_merged_df.loc[np.r_[
    #         0:total_fractions], :]

    # elif (my_merged_df.shape[0] > 40) and (len(my_merged_df['Sample barcode'].unique().tolist()) == 2):
    #     my_merged_df = my_merged_df.loc[np.r_[
    #         0:total_fractions, 40:(total_fractions+40)], :]

    # else:
    #     # get the sip plates with unexpected number of samples or fractions
    #     sip_plate = []
    #     sip_plate = my_merged_df['Plate barcode'].unique().tolist()

    #     print(
    #         f'\n\nToo many sample barcodes or fractions in plate {sip_plate}\n\n')
    #     sys.exit()
    
    
    # select only wells where fractions were collected
    my_merged_df = my_merged_df[my_merged_df['Fraction #'] <= total_fractions]
    
    if (my_merged_df.shape[0] < total_fractions) or (len(my_merged_df['Sample barcode'].unique().tolist()) == 0):
        print(f'\n Failed to successfully add sequin info to samples {sample_list}. Aborting process\n\n')
        
        sys.exit()

    elif (my_merged_df.shape[0] > (2*total_fractions)) or (len(my_merged_df['Sample barcode'].unique().tolist()) > 2):
        print('\n Too many fractions or plates when trying to add sequin info to samples {sample_list}.  Aborting process\n\n')
        
        sys.exit()

    return my_merged_df
##########################
##########################


##########################
##########################
def makeTransferFiles(my_merged_df, plateid, well_list_96w):
    trans_df = my_merged_df[['Plate barcode', 'Sample barcode', 'Well',
                             'actual_sequin_mass_(ng)', 'sequin_working_stock_conc_(pg/uL)', 'sequin_vol_(uL)']].copy()

    # Rename the column header to reflect that values are actually in pg, not ng
    trans_df = trans_df.rename(columns={'actual_sequin_mass_(ng)': 'actual_sequin_mass_(pg)'})

    # create dict here key is well position and value in numerical order column-wise
    # will use later for sorting df by well position
    well_dict = dict((k, v) for v, k in enumerate(well_list_96w))

    # add column with column-wise rank order of well positions
    trans_df['order'] = trans_df['Well'].map(well_dict)

    # # sort df first by sequin stock concentration, then by column-wise by well
    # trans_df = trans_df.sort_values(
    #     by=['sequin_working_stock_conc_(pg/uL)', 'order'], ascending=[False, True])

    # sort df column-wise by well
    trans_df = trans_df.sort_values(
        by=['order'], ascending=[True])

    trans_df.drop(columns=['order'], inplace=True)

    # write .csv transfer file for hamilton
    trans_df.to_csv(SEQUIN_DIR / f'sequin_{plateid}_{date}.csv', index=False)
    return
##########################
##########################


##########################
##########################
def summarizeSequins(all_plate_df):
    # make new df with subset of column in all_plate_df
    sequin_df = all_plate_df[[
        'Plate barcode', 'Sample barcode', 'Well', 'sample_mass_(ng)', 'sequin_stock', 'sequin_working_stock_conc_(pg/uL)', 'actual_sequin_mass_(ng)', 'sequin_vol_(uL)']].copy()

    # get list of plates that were processed, and sort list
    plate_list = list(sequin_df['Plate barcode'].unique())

    plate_list.sort()

    # estimate total volume sequin working stocks needed
    # multiply by 1.05 to add 5% overage, and add 10ul for dead volume
    hi_total_sequin_vol = round(
        1.05 * sequin_df.loc[sequin_df['sequin_stock'] == 'HI', 'sequin_vol_(uL)'].sum(), 0)+10

    lo_total_sequin_vol = round(
        1.05 * sequin_df.loc[sequin_df['sequin_stock'] == 'lo', 'sequin_vol_(uL)'].sum(), 0)+10

    print(
        f'\n\n\n You will need a total HIGH sequin volume of {hi_total_sequin_vol} uL for these plates\n')

    print(
        f' You will need a total LOW sequin volume of {lo_total_sequin_vol} uL for these plates\n\n')

    # calculate the total ng's recovered in pre PEG fractions
    total_ng_df = sequin_df.groupby('Sample barcode')[
        'sample_mass_(ng)'].sum().round(0)

    outfile_path = f"Summary_sequin_volumes_{date}.txt"

    # create text file with summary of total volume  of sequin stock needed
    # out = open(SEQUIN_DIR / f"sequin_volume_summary_{date}.txt", "w")
    out = open(SEQUIN_DIR / outfile_path, "w")
    out.write(f"Total volume HIGH sequin stock:\t{hi_total_sequin_vol}\n")
    out.write(f"Total volume low sequin stock:\t{lo_total_sequin_vol}\n\n")
    out.write("SIP plates included:\n\n")
    # write names of all plates processed in this batch
    for pl in plate_list:
        out.write(f"{pl}\n")

    out.write("\n\n")

    out.close()

    total_ng_df.to_csv(SEQUIN_DIR / outfile_path, header=True,
                       index=True, sep='\t', mode='a')

    return sequin_df
##########################
##########################


##########################
##########################
def updateDensityFile(my_merged_df, my_density_emptyrow_df, plateid):

    seqmass_df = my_merged_df[['Sample barcode',
                               'Well Pos', 'actual_sequin_mass_(ng)']].copy()

    # merge df from density file with df containing actual sequin mass added to each fraction
    dens_df = my_density_emptyrow_df.merge(seqmass_df, how='outer', left_on=['Sample barcode', 'Well Pos'],
                                           right_on=['Sample barcode', 'Well Pos'])

    # make sure merging worked and there were not misalignments.
    if (dens_df.shape[0] != my_density_emptyrow_df.shape[0]):
        print('\nProblem adding sequin mass to density .xlsx file. aborting\n\n')
        sys.exit()

    dens_df = dens_df.fillna('')

    # reduce dens_df to a single column after successfully aligning and merging with density df
    dens_df = dens_df[['actual_sequin_mass_(ng)']]

    # create variable with density file name
    fn = (f'{plateid}.xlsx')
    full_path = BASE_DIR / fn

    # save copy of unmodified density .xlsx files to a sub folder
    shutil.copy(full_path, DENSITY_DIR / f'{plateid}_{date}.xlsx')

    # # one option for writing sequin mass to density .xlsx file
    # book = openpyxl.load_workbook(full_path)

    # writer = pd.ExcelWriter(full_path, engine='openpyxl')

    # writer.book = book

    # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # try:
    #     dens_df.to_excel(writer, sheet_name='updated', header=None, index=False,
    #                      startcol=6, startrow=1)

    # except:
    #     print('\nProblem writing to density .xlsx with sequin mass. Aborting\n')
    #     sys.exit()

    # finally:
    #     writer.close()

    #     book.close()

    # Use openpyxl directly for more reliable Excel writing across pandas versions
    from openpyxl import load_workbook
    
    # Load the existing workbook
    wb = load_workbook(full_path)
    ws = wb['updated']
    
    # Create a mapping from well position to sequin value from the merged data
    well_to_sequin = {}
    for _, row in my_merged_df.iterrows():
        well = row['Well Pos']
        sequin_value = row['actual_sequin_mass_(ng)']
        well_to_sequin[well] = sequin_value
    
    # Write sequin values to correct Excel rows based on well position matching
    for i, row in my_density_emptyrow_df.iterrows():
        well = row['Well Pos']
        if well in well_to_sequin:
            sequin_value = well_to_sequin[well]
            ws.cell(row=i+2, column=7, value=sequin_value)  # +2 for 1-based indexing and header row, column 7 is G
    
    # Save the workbook
    wb.save(full_path)
    wb.close()


    

    
    # rename density .xlsx file with '~$' appened to start of file name
    # this simulates Microsoft name format for temporary files, and changing
    # the name to this temporary format is necessary so excel on Mac can open
    # the .xlsx files the first time without granting access file by file
    # the name will be changed back to the original format when donde
    temp_name = BASE_DIR / ("~$"+fn)
    
    os.rename(full_path, temp_name)
    
    
    
    # open modified density .xlsx file in excel, save, and close
    # this resovles an error where .xlsx formulas stop working, e.g. density calculations
    # valres are not imported into pandas df in down stream python merging script run after
    # this script
    app = xw.App(visible=False)
    # book = app.books.open(full_path)
    
    book = app.books.open(temp_name)
    book.save()
    app.kill()
    
    #rename density .xlsx file by removing the '~$' from the start of the file name
    os.rename(temp_name, full_path)

    # another option for writing sequin mass to density .xlsx files
    # # load workbook
    # app = xw.App(visible=False)
    # wb = xw.Book(f'{plateid}.xlsx')
    # ws = wb.sheets['updated']

    # # #Update workbook at specified range
    # ws.range('G2').options(index=False).value = dens_df

    # # #Close workbook
    # wb.save()
    # wb.close()
    # app.quit()

    return
##########################
##########################


# ############################
# ############################
# def updateAllPlate(my_result_df, my_project_df):

#     # make group_dict dictionary where key is ITS sample ID and value is replicate group
#     group_dict = dict(zip(my_project_df.ITS_sample_id,
#                       my_project_df.Replicate_Group))

#     # make group_dict dictionary where key is ITS sample ID and value is replicate group
#     name_dict = dict(zip(my_project_df.ITS_sample_id,
#                      my_project_df.Sample_Name))

#     # add column using group_dict
#     my_result_df['Replicate_Group'] = my_result_df['Sample barcode'].map(
#         group_dict)

#     # add column of fraction sample name by concat parent sample name with the fraction #
#     my_result_df['Fraction_sample_name'] = my_result_df['Sample barcode'].map(
#         name_dict) + "_" + my_result_df['Fraction #'].astype(str)

#     # make sure replicate group and fraction sample ids were assigned to all fractions
#     if (my_result_df['Replicate_Group'].isnull().values.any()):
#         print("\n\n")
#         print(
#             '\n\n Could not assign Replciate_Group to at least one sample. Aborting. Check error file. \n\n')

#         my_result_df.to_csv('error.desnity.volume.merge.csv', index=False)
#         sys.exit()

#     elif (my_result_df['Fraction_sample_name'].isnull().values.any()):
#         print("\n\n")
#         print(
#             '\n\n Could not assign Fraction_sample_name to at least one sample. Aborting. Check error file. \n\n')
#         my_result_df.to_csv('error.desnity.volume.merge.csv', index=False)
#         sys.exit()

#     # get a list of columns
#     cols = list(my_result_df)
#     # move the column to head of list using index, pop and insert
#     cols.insert(0, cols.pop(cols.index('Replicate_Group')))

#     # move the column to head of list using index, pop and insert
#     cols.insert(0, cols.pop(cols.index('Fraction_sample_name')))

#     # use ix to reorder
#     my_result_df = my_result_df.loc[:, cols]

#     return my_result_df
# ############################
# ############################


# ##########################
# ##########################
# def makeDNAvsDensityPlots(all_plate_df, total_fractions):

#     # get list of unique group names in project
#     groups = sorted(all_plate_df['Replicate_Group'].unique().tolist())

#     # empty list to later hold names of pdfs for each group figure
#     # this will be used ot merge multiple pdfs at end of script
#     pdf_files = []

#     # loop through list of groups
#     for g in groups:

#         # get list of samples in group
#         samples = sorted(all_plate_df[all_plate_df['Replicate_Group'] == g]
#                          ['Sample barcode'].unique().tolist())

#         # creation matliplot object
#         f = plt.figure()

#         # loop through samples and add line to plt.plot
#         for s in samples:
#             x = all_plate_df[(all_plate_df['Sample barcode'] == s) & (
#                 all_plate_df['Fraction #'].between(1, total_fractions))]['Density']
#             y = all_plate_df[(all_plate_df['Sample barcode']
#                               == s) & (all_plate_df['Fraction #'].between(1, total_fractions))]['[Concentration]']

#             # x = lib_df[(lib_df['Sample barcode'] == s) & (
#             #     lib_df['Density (g/mL)'].between(1.669, 1.781))]['Density (g/mL)']
#             # y = lib_df[(lib_df['Sample barcode']
#             #            == s) & (lib_df['Density (g/mL)'].between(1.669, 1.781))]['DNA Concentration (ng/uL)']

#             # x = all_plate_df[(all_plate_df['Sample barcode'] == s) & (
#             #     all_plate_df['Make_Lib'] == 1)]['Density (g/mL)']
#             # y = all_plate_df[(all_plate_df['Sample barcode']
#             #                   == s) & (all_plate_df['Make_Lib'] == 1)]['DNA Concentration (ng/uL)']

#             # make new column the concats sample barcode with fraction_sample_name
#             all_plate_df['combo_name'] = all_plate_df['Sample barcode'] + \
#                 '-'+all_plate_df['Fraction_sample_name']

#             # extract user provided sample name by removing the suffix "_fraction#" from fraction name
#             name_list = all_plate_df[all_plate_df['Sample barcode'] ==
#                                      s]['combo_name'].unique().tolist()

#             # remove "_fraction#' from sample name
#             ch = '_'
#             parts = name_list[0].split(ch)
#             parts.pop()
#             name = ch.join(parts)

#             # add plot of current sample
#             plt.plot(x, y, label=name, marker='o')

#         # add plot title, axes labels, and legend to group plot
#         plt.xlabel("Density (g/mL)")
#         plt.ylabel("DNA Concentration (ng/uL)")
#         plt.title(f'Replicate Group: {g}')
#         plt.legend(fontsize='x-small')
#         # plt.show()
#         f.savefig(f'UserName_DNAvsDensity_{g}.pdf',
#                   format="pdf", bbox_inches="tight")

#         # add name of group figure pdf
#         pdf_files.append(f'UserName_DNAvsDensity_{g}.pdf')

#         plt.close(f)

#     # Create and instance of PdfFileMerger() class
#     merger = PdfFileMerger()

#     # loop through group figure pdfs and merge them into one
#     for pdf in pdf_files:
#         # Append PDF files
#         merger.append(pdf)
#         # # delete indidvidual group figure pdf
#         # os.remove(pdf)

#     # Write out the merged PDF
#     merger.write(PROJECT_DIR / f"Pre_PEG_DNAvsDensity_plots_{date}.pdf")
#     merger.close()

#     for pdf in pdf_files:
#         # delete indidvidual group figure pdf
#         os.remove(pdf)

#     return
# ##########################
# ##########################


##########################
##########################
###    MAIN Program    ###
##########################
##########################

# get current date and time, will add some file names
date = datetime.now().strftime("%Y_%m_%d-Time%H-%M-%S")


# set currign working directory and create a 'sequins' folder if it doesn't exist already


PROJECT_DIR = Path.cwd()

BASE_DIR = PROJECT_DIR / "3_merge_density_vol_conc_files"

SEQUIN_DIR = BASE_DIR / "sequins"

SEQUIN_DIR.mkdir(parents=True, exist_ok=True)

DENSITY_DIR = BASE_DIR / "original_unmodified_density_files"

DENSITY_DIR.mkdir(parents=True, exist_ok=True)

# PREDNA_DIR = BASE_DIR / "DNAconc_files_pre_PEG"

# PREDNA_DIR.mkdir(parents=True, exist_ok=True)

# get current directory
current_directory = os.getcwd()

subdirectory_name = "3_merge_density_vol_conc_files"

# create path to subdirectory where density, volume, and dna conc files are located
dirname = os.path.join(current_directory, subdirectory_name)

# file prefix for DNA conc files
prefix = ('pre')

# get list of matched sets of density, dna conc, and volume files
dna_files = getMatchedFiles(dirname, prefix)

total_fractions, percent_sequin, cscl_vol, hi_sequin_working_conc, lo_sequin_working_conc, min_trans_vol, max_trans_vol = enterFractionMetadata()

# create empty list to hold all plate ids found
list_merged_plates = []

# create empty df to hold results for all plates
all_plate_df = pd.DataFrame()

# loop through all dna conc plates and merge with density and volume files
# and calculate amount of sequins to add to each fraction
for dna in dna_files:

    # generate df of DNA conc values
    DNA_conc_df = getConc(dirname, dna)

    # get plate name by removing DNA conc file prefix
    plateid = dna.replace(prefix, "")
    plateid = plateid.replace(".txt", "")

    # creat df's from density and volume files
    volume_df = getVolumes(dirname, plateid)
    density_df, density_emptyrow_df = getDensity(dirname, plateid)

    # merge dna conc, density, and volume df's
    merged_df, list_merged_plates = getMergedPlates(
        DNA_conc_df, volume_df, density_df, plateid, list_merged_plates)

    # calculate sequin mass and volume additions for each fraction
    merged_df = addSequinVol(merged_df, total_fractions, percent_sequin, cscl_vol,
                             hi_sequin_working_conc, lo_sequin_working_conc, min_trans_vol, max_trans_vol)

    # make hamilton transfer files for adding sequins
    # and make sequin_df
    makeTransferFiles(merged_df, plateid, well_list_96w)

    # updated density files with sequin mass added
    updateDensityFile(merged_df, density_emptyrow_df, plateid)

    # concat merged_df from current dna plate with master df of all dna plates
    all_plate_df = pd.concat([all_plate_df, merged_df],
                             axis=0, ignore_index=True)

# make summary of sequins batch
sequin_df = summarizeSequins(all_plate_df)

# # loop through dna conc file with "pre" prefix again, and move them to a subfolder
# for dna in dna_files:
#     # move DNA conc file to a subfolder for files with prefix
#     shutil.move(BASE_DIR / dna, PREDNA_DIR / dna)


# # read project_database.csv into  a dataframe
# project_df = pd.read_csv(PROJECT_DIR / 'project_database.csv', header=0, usecols=['Sample_Name', 'Replicate_Group', 'ITS_sample_id'], converters={
#     'ITS_sample_id': str})

# # add user sample name and replicate group to all_plate_df
# all_plate_df = updateAllPlate(all_plate_df, project_df)

# # make pdf plots of density vs DNA conc
# makeDNAvsDensityPlots(all_plate_df, total_fractions)

# Create success marker to indicate script completed successfully
from pathlib import Path
status_dir = Path.cwd() / ".workflow_status"
status_dir.mkdir(exist_ok=True)
success_file = status_dir / "calcSequinAddition.success"
success_file.touch()
