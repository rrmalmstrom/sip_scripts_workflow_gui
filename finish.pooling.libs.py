#!/usr/bin/env python3

# USAGE:   python finish.pooling.libs.py <pool.creation.file.xls>

import pandas as pd
import numpy as np
import sys
import os
import xlwings as xw
import shutil
from datetime import datetime
from pathlib import Path
import openpyxl
import string
import random
import math
from sqlalchemy import create_engine



##########################
##########################
def readSQLdb(my_prjct_dir):

    # path to sqlite db lib_info_submitted_to_clarity.db
    sql_db_path = f'{my_prjct_dir}/lib_info_submitted_to_clarity.db'

    # create sqlalchemy engine
    engine = create_engine(f'sqlite:///{sql_db_path}') 

    # define sql query
    query = "SELECT * FROM lib_info_submitted_to_clarity"
    
    # import sql db into pandas df
    sql_df = pd.read_sql(query, engine)
    
    sql_df['Sample Barcode'] = sql_df['Sample Barcode'].astype('str')
    
    sql_df['Fraction #'] = sql_df['Fraction #'].astype('int')
     
    engine.dispose()

    return sql_df
##########################
##########################


##########################
##########################
def findPoolFile(crnt_dir):
    
    clarity_pool_prep_files = []

    # search through current directory for Clarity Pool Prep file
    # format is PoolingPrep_XX-XXXXXXXX.xlsx
    for file in os.listdir(crnt_dir):
        if file.startswith('PoolCreation_'):
            
            # add any files matching pattern to list
            clarity_pool_prep_files.append(file)
          
    # abort script if no pool creation files are found or 
    # if >1 pool creation file found        
    if len(clarity_pool_prep_files) == 0:
        print('\nCould not find Clarity pool prep file, e.g. PoolCreation_XX-XXXXX.xslx   Aborting script\n\n')
        sys.exit()
    elif len(clarity_pool_prep_files) > 1:
        print(f'\nMultiple Clarity pool creation files found\n\n{clarity_pool_prep_files}\n\nAborting script\n\n')
        sys.exit()

    pool_file= clarity_pool_prep_files[0]

    # return Clarity pool creation file name
    return pool_file
##########################
##########################


##########################
##########################
def fixExcelFile(pool_file):
    
    # rename Pool prep .xlsx file with '~$' appened to start of file name
    # this simulates Microsoft name format for temporary files, and changing
    # the name to this temporary format is necessary so excel on Mac can open
    # the .xlsx files the first time without granting access file by file
    # the name will be changed back to the original format when donde
    
    # Use full path for file operations
    pool_file_path = Path(crnt_dir) / pool_file
    temp_name_path = Path(crnt_dir) / ("~$" + pool_file)
    
    os.rename(pool_file_path, temp_name_path)
    
    
    
    # open pool prep .xlsx file in excel, save, and close
    # this resovles an error where .xlsx formulas stop working, e.g.
    # values are not imported into pandas
    app = xw.App(visible=False)
    book = app.books.open(str(temp_name_path))
    book.save()
    app.kill()
    
    #rename density .xlsx file by removing the '~$' from the start of the file name
    os.rename(temp_name_path, pool_file_path)
    
##########################
##########################


##########################
##########################
def addPoolClarityID(lib_df, pool_file):

    pool_file_path = Path(crnt_dir) / pool_file
    
    ####### MUST REACTIVATE THIS SECTION AND COMMENT OUT REFERENCE TO pool_df.csv #########

    # import library creation .xls file downloaded from clarity queue
    pool_df = pd.read_excel(pool_file_path, sheet_name="Lab", header=0, usecols=[
        " qPCR'd Sample Name", "Sample Container Barcode", "Source Position", "Pool Name", "Pool Container Barcode"])

   
    pool_df.rename(columns={" qPCR'd Sample Name": "clar_Lib_id", "Sample Container Barcode": "clar_Source Container Barcode",
                    "Source Position": "clar_Source Position", "Pool Name": "Pool_id", "Pool Container Barcode": "Pool_LIMS"}, inplace=True)

    # ############## THIS LINE READING FROM POOL_DF.CSV MUST BE DEACTIVATED WHEN WORKING WITH REAL SAMPLES ############## 
    # pool_df = pd.read_csv('pool_df.csv')

    # confirm that there were no missing values or  mismatches between lib_df angd pool_df
    if pool_df.shape[0] != lib_df.shape[0]:
        print("\nThere was a mismatch between lib_info_submitted df and pool creation .xls sheets.  Aborting process\n")
        sys.exit()

    # add clarity ids to lib_info
    merged_df = lib_df.merge(pool_df, how='outer', left_on=['Library Name', 'Clarity_Lib_Plate_ID', 'Pool_source_well'], right_on=[
                             'clar_Lib_id', 'clar_Source Container Barcode', 'clar_Source Position'])

    # drop unnecesasry columns
    merged_df.drop(['clar_Lib_id', 'clar_Source Container Barcode',
                   'clar_Source Position'], inplace=True, axis=1)

    # confirm that there were no missing values or  mismatches between lib_df angd pool_df
    if merged_df.shape[0] != lib_df.shape[0]:
        print("\nThere was a mismatch between lib_info_submitted df and pool creation .xls sheets.  Aborting process\n")
        sys.exit()
    elif merged_df['Pool_id'].isnull().values.any():
        print("\nThere was a mismatch between lib_info_submitted df and pool creation .xls sheets.  Aborting process\n")
        sys.exit()

    # update the destination tube info as concat of pool# and clarity plate id
    merged_df['Destination_Tube_Name'] = merged_df['Pool_number'].astype(
        str) + "a_" + merged_df['Pool_id']

    # merged_df['Destination_Tube_Barcode'] = merged_df['Destination_Tube_Name']

    merged_df['Destination_Tube_Barcode'] = merged_df['Pool_LIMS']

    return merged_df
##########################
##########################


##########################
##########################
def addPippinInfo(my_lib_df):

    # make list of pools then sort by pool #
    pool_list = my_lib_df['Pool_number'].unique().tolist()

    pool_list.sort()

    # generate unique 6 digit ID destination/lib creation plate, first digit is a letter
    destbc = random.choice(string.ascii_uppercase)

    destbc = destbc + \
        "".join(random.choice(string.digits+string.ascii_uppercase)
                for _ in range(5))

    # create empty dicts for linkin pool# with lane# and cassette#
    lane_dict = {}

    cassette_dict = {}
   
    
    # create tuple of tuples indicating pairs of pippin lanes
    # tuple (1,3) indicates a pool should be loaded into lanes 1 and 3
    lane_pairs = ((1,3),(5,7),(9,11),(2,4),(6,8))
    
    
    # list through all pool# and create two dicts where key is pool# and
    # value is either the pippin lane# (lane_dict) or pippin cassette barcode
    # (cassette_dict) using modulo 10 instead of 12 because lanes 11 and 12 are used
    # for library standard positive controls and ladder
    for i, p in enumerate(pool_list):
        if (i+1) % 5 == 0:
            lane_dict[p] = lane_pairs[4]
        else:
            pair_position = ((i+1) % 5) - 1
            lane_dict[p] = lane_pairs[pair_position]

        cassette_dict[p] = destbc + "-" + str((math.floor((i)/5))+1)

    # add column with Pippin lane # to df, but lane is a tuple because
    # pool is split into 2 lanes indicated in lane_pairs
    my_lib_df['Pippin_Lane'] = my_lib_df['Pool_number'].map(lane_dict)
    
    # df[['b1', 'b2']] = pd.DataFrame(df['b'].tolist(), index=df.index)
    my_lib_df[['1st_Pippin_lane','2nd_Pippin_lane']] = pd.DataFrame(my_lib_df['Pippin_Lane'].tolist(),index=my_lib_df.index)
    
    # drop column containing lane tuple
    my_lib_df.drop(['Pippin_Lane'], axis=1, inplace=True)


    # add column with Pippin cassette barcode to df
    my_lib_df['Pippin_Cassette'] = my_lib_df['Pool_number'].map(cassette_dict)

    return my_lib_df
##########################
##########################

##########################
##########################
def fillPoolCreationSheet(lib_df, pool_file):

    pool_file_path = Path(crnt_dir) / pool_file

    clarity_df = pd.read_excel(
        pool_file_path, header=0, sheet_name='Lab', usecols="A:AE")

    # copy lib percentage values to column with missing values
    clarity_df['Library Actual Percentage with SOF'] = clarity_df['Library Percentage with SOF']

    # fill required columns
    clarity_df['Pool Lab Process Result'] = "Pass"

    # this conc is a guess and will later be updated when pool is run through qPCR queue
    clarity_df['Pool Concentration pM'] = 1500

    # fill in other columns normally used for pooling with dummy data
    clarity_df['Conc (pM) '] = 5000

    clarity_df['Library Working Concentration for Pooling'] = 5000

    clarity_df['ul of Undiluted or 1:10 Library'] = 3.0

    clarity_df['Vol TE  to Bring Library up to 10ul'] = 7.0

    clarity_df['Target Pool Concentration pM'] = 1500

    clarity_df['Actual Volume Used'] = 3.0

    clarity_df['Repool Volume Present?'] = 'Y'

    clarity_df['Destination Labware'] = 'Pool Tube Rack'

    # replace Nan with ''
    clarity_df = clarity_df.fillna('')

    # read in dataframe from Lab summary table sheet
    summary_df = pd.read_excel(
        pool_file_path, header=0, sheet_name='Lab Summary Table', usecols="A:E")

    # add dummy values to dataframe
    summary_df['Final Target Volume'] = 999

    summary_df['Target Concentration'] = 1500

    summary_df['Pooled Volume'] = 999

    summary_df['Average Fragment Size'] = 555

    # make a copy of the pool creattion file, and fill new values into this copy
    copy_pool_file = Path(crnt_dir) / ('autofilled_'+ pool_file)

    shutil.copyfile(pool_file_path, copy_pool_file)

    # load pool creation xlsx and delete existing values in first 2000 rows
    wb = openpyxl.load_workbook(copy_pool_file)

    sh1 = wb['Lab']

    sh1.delete_rows(2, 2000)

    # loop through clarity df and write info into Lab sheet in pool creation .xlsx file
    for row_num, row in clarity_df.iterrows():
        for col_num, col in enumerate(row):
            cellref = sh1.cell(row=(2+row_num), column=(1+col_num))
            cellref.value = col

    sh2 = wb['Lab Summary Table']

    sh2.delete_rows(2, 2000)

    # loop through summary df and write info into Lab stummary table sheet in pool creation .xlsx file
    for row_num, row in summary_df.iterrows():
        for col_num, col in enumerate(row):
            cellref = sh2.cell(row=(2+row_num), column=(1+col_num))
            cellref.value = col

    wb.save(copy_pool_file)

    wb.close()

    return
##########################
##########################


##########################
##########################
def makePoolTransferFiles(my_lib_df):

    # get list of all pool numbers
    pool_list = my_lib_df['Pool_number'].unique().tolist()

    # loop through each pool number and make a transfer file for the Hamilton
    for p in pool_list:

        # make temporary df with only libs with same pool number
        tmp_df = my_lib_df.loc[my_lib_df['Pool_number'] == p].copy()


        tmp_df = tmp_df[['Pool_transfer_plate',
                        'Pool_source_well', 'Pool_ACTUAL_transfer_volume_(uL)', 'Destination_Tube_Name']]

        # the barcode on the 15mL tube used for pooling is teh same as the pool name
        # e.g. 1_NUPUS
        # the clarity container barcode will be used only on tube that collects
        # DNA after pippin size selection
        tmp_df['Destination_Tube_Barcode'] = tmp_df['Destination_Tube_Name']

        tmp_df['Source_Barcode'] = tmp_df['Pool_transfer_plate']

        tmp_df = tmp_df.rename(
            columns={'Pool_transfer_plate': 'Source_Name', 'Pool_source_well': 'Source_Well', 'Pool_ACTUAL_transfer_volume_(uL)': 'Transfer_Volume'})

        # Destination_Tube_Name, Destination_Tube_Name

        final_col_list = ['Source_Name', 'Source_Barcode',
                          'Source_Well', 'Transfer_Volume', 'Destination_Tube_Name', 'Destination_Tube_Barcode']

        tmp_df = tmp_df.reindex(columns=final_col_list)

        tmp_df['Myindex'] = tmp_df.index

        tmp_df = tmp_df.sort_values(by=['Source_Name', 'Myindex'])

        # reset index so that later concat with bcode_df starts on index == 0
        tmp_df.reset_index(drop=True, inplace=True)

        tmp_df.drop(['Myindex'], axis=1, inplace=True)

        tmp_df.to_csv(ATTEMPT_DIR / f'Pool_{p}_transfer_file.csv', index=False, header=True)
        
        # tmp_df.to_csv(f'Pool_{p}_transfer_file.csv', index=False, header=True)

    return
##########################
##########################


##########################
##########################
def makeTubeBarcodeFiles(my_lib_df):
    
    pippin_df = my_lib_df[['Destination_Tube_Name',
                           'Destination_Tube_Barcode', 'Pippin_Cassette', '1st_Pippin_lane','2nd_Pippin_lane']].copy()

    # remove duplicate rows, then sort by destination_tube_name, and reset the index
    pippin_df = pippin_df.drop_duplicates()
    
    pippin_df.sort_values(by=['Destination_Tube_Name'], ascending=True, inplace=True)


    pippin_df.reset_index(drop=True, inplace=True)
    
    pippin_df['Dest_Tube_Size_Selected'] = pippin_df['Destination_Tube_Name']
    
    pippin_df['Dest_Tube_Size_Selected'] = pippin_df['Dest_Tube_Size_Selected'].str.replace('a_','-1_')


    # make dict where key is pool name and value is pool container id
    pool_dict = dict(zip(pippin_df.Destination_Tube_Name,
                      pippin_df.Dest_Tube_Size_Selected))

    # get list of unique pool names
    my_dest_list = my_lib_df['Destination_Tube_Name'].unique().tolist()
    
    my_dest_list.sort()


    # this was older format for bartender templates.  The newer version below changes "/" to "\"
    # in the path to the template files  AF="*"
    # x = '%BTW% /AF="//bartender/shared/templates/JGI_Label_BCode5_Rex.btw" /D="%Trigger File Name%" /PRN="bcode5" /R=3 /P /DD\r\n%END%\r\n\r\n\r\n'

    x = '%BTW% /AF="\\\\bartender\shared\\templates\JGI_Label_BCode5_Rex.btw" /D="%Trigger File Name%" /PRN="bcode5" /R=3 /P /DD\r\n%END%\r\n\r\n\r\n'


    bc_file = open(ATTEMPT_DIR /"pooling_TUBES_barcode_labels.txt", "w")
    
    # bc_file = open("pooling_TUBES_barcode_labels.txt", "w")

    bc_file.writelines(x)

    # add barcodes of library destination plates, dna source plates, and buffer plate
    for p in my_dest_list:
        tube_name = p.split('_')

        # make multiple copies of barcode labels
        for z in range(3):
            # bc_file.writelines(f'{p},{p}\n')
            bc_file.writelines(
                f'{p},{tube_name[0]},{tube_name[1]}\r\n')



        size_tubename = pool_dict[p].split('_')
        
        bc_file.writelines(
            f'{pool_dict[p]},{size_tubename[0]},{size_tubename[1]}\r\n')


        
        # add a row of blank labels
        bc_file.writelines(',,\r\n')

    bc_file.close()

    return pippin_df
##########################
##########################


##########################
##########################
def FAplatePoolQC(my_pippin_list):

    # list will be refernced to convert row number into row letter
    row_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

    # create empty dict where key will be pool name and value
    # will be FA well position
    well_dict = {}

    # loop through list of pool names
    for i, p in enumerate(my_pippin_list):

        # set row number as floor after dividing by 10
        r = math.floor((i)/10)
    

        # set column number as modulo of 10
        # use modulo 10 so that no pool is assigned
        # to columns 11 or 12. They ar reserved for
        # control and ladder in the FA plate
        c = ((i+1) % 10)

        if c == 0:
            c = 10

        well_dict[p] = row_list[r]+str(c)

    return well_dict
##########################
##########################


##########################
##########################
def makePippinTransferFiles(my_pippin_df):

    my_pippin_df.rename(columns={'Destination_Tube_Name': 'Pool_Name',
                                 'Destination_Tube_Barcode': 'Pool_Barcode'}, inplace=True)

    # add pre-determined transfer volumes
    # my_pippin_df['Sample_volume_(uL)'] = 22

    # my_pippin_df['Marker_volume_(uL)'] = 5.5
    
    my_pippin_df['Sample_volume_(uL)'] = 44

    my_pippin_df['Marker_volume_(uL)'] = 11

    my_pippin_df['Load_volume_(uL)'] = 25

    my_pippin_df['Recover_volume_(uL)'] = 30

    # create list with unique pool names
    pippin_list = my_pippin_df['Pool_Name'].unique().tolist()

    # # sort list of pool names
    # pippin_list.sort()

    # call function to get dict where key is pool name and value is
    # well position if FA plate for pool QC
    fa_well_dict = FAplatePoolQC(pippin_list)

    # generate unique 6 digit ID destination/lib creation plate, first digit is a letter
    fabc = random.choice(string.ascii_uppercase)

    fabc = fabc + \
        "".join(random.choice(string.digits+string.ascii_uppercase)
                for _ in range(4))
    
    # add -FA to FA plate name to help distinguish it from pippin cassette
    fabc = fabc + '-FA1'

    my_pippin_df['FA_plate_barcode'] = fabc

    # add column with FA well position
    my_pippin_df['FA_well'] = my_pippin_df['Pool_Name'].map(fa_well_dict)

    my_pippin_df['FA_transfer_vol_(uL)'] = 2.4

    # create pippin transfer file
    my_pippin_df.to_csv(ATTEMPT_DIR /
        'PIPPIN_load_unload_transfer_file.csv', index=False, header=True)
    
    # # create pippin transfer file
    # my_pippin_df.to_csv(
    #     'PIPPIN_load_unload_transfer_file.csv', index=False, header=True)

    # create list with unique FA plate barcodes
    dest_list = my_pippin_df['FA_plate_barcode'].unique().tolist()

    dest_list.sort()

    return my_pippin_df, dest_list
##########################
##########################


##########################
##########################
def makePippinBarcodeFile(my_pippin_df):

    # get list of all pool numbers
    pippin_list = my_pippin_df['Pippin_Cassette'].unique().tolist()

    fa_list = my_pippin_df['FA_plate_barcode'].unique().tolist()


    # this was older format for bartender templates.  The newer version below changes "/" to "\"
    # in the path to the template files  AF="*"
    # x = '%BTW% /AF="//BARTENDER/shared/templates/ECHO_BCode8.btw" /D="%Trigger File Name%" /PRN="bcode8" /R=3 /P /DD\r\n\r\n%END%\r\n\r\n\r\n'
    
    # add info to start of barcode print file indicating the template and printer to use
    x = '%BTW% /AF="\\\BARTENDER\shared\\templates\ECHO_BCode8.btw" /D="%Trigger File Name%" /PRN="bcode8" /R=3 /P /DD\r\n\r\n%END%\r\n\r\n\r\n'
    
    bc_file = open(ATTEMPT_DIR /"PIPPIN_CASSETTE_FA_PLATE_barcode_labels.txt", "w")
    

    bc_file.writelines(x)

    for p in pippin_list:
        bc_file.writelines(f'{p},"pip.casset {p}"\r\n')


    for fa in fa_list:
        bc_file.writelines(f'{fa},"FA.plate {fa}"\r\n')


    bc_file.close()

    return
##########################
##########################


##########################
##########################
# this verseion fo function assuming pool will be split into 1 pippin lane
def getWellList(my_FA_df):

    # list will be refernced to convert row number into row letter
    row_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

    well_list = []

    num_fa_wells = len(my_FA_df['Dest_Tube_Size_Selected'].unique().tolist())

    num_rows = math.ceil(num_fa_wells/10)

    for i in range(1, (12*num_rows)+1):

        # set row number as floor after dividing by 12
        r = math.floor((i-1)/12)

        # set column number as modulo of 12
        c = ((i) % 12)

        if c == 0:
            well_list.append(row_list[r]+'12')

        else:
            well_list.append(row_list[r]+str(c))

    return well_list, num_rows
##########################
##########################



##########################
##########################
def makeFAinputFiles(my_pippin_df, my_dest_list):

    # list will be refernced to convert row number into row letter
    row_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

    FA_df = my_pippin_df[['Dest_Tube_Size_Selected',
                          'FA_plate_barcode', 'FA_well']].copy()

    FA_df['FA_plate_barcode'] = FA_df['FA_plate_barcode'].astype(str)

    FA_df['Dest_Tube_Size_Selected'] = FA_df['Dest_Tube_Size_Selected'].astype(str)

    FA_df['FA_well'] = FA_df['FA_well'].astype(str)

    FA_df['name'] = FA_df[['FA_plate_barcode', 'Dest_Tube_Size_Selected', 'FA_well']].agg(
        '.'.join, axis=1)

    my_well_list, my_num_rows = getWellList(FA_df)

    # drop unecessary columns so only have df with 'Destination_Well', 'Destination_ID', and 'name'
    FA_df.drop(['Dest_Tube_Size_Selected'], inplace=True, axis=1)

    tmp_fa_df = pd.DataFrame(my_well_list)

    tmp_fa_df.columns = ["Well"]

    tmp_fa_df = tmp_fa_df.merge(FA_df, how='outer', left_on=[
                                'Well'], right_on=['FA_well'])

    tmp_fa_df.index = range(1, tmp_fa_df.shape[0]+1)

    fabc = tmp_fa_df['FA_plate_barcode'].iloc[0]

    tmp_fa_df.drop(['FA_plate_barcode', 'FA_well'], inplace=True, axis=1)

    tmp_fa_df['name'] = tmp_fa_df['name'].fillna('empty_well')

    start = 0
    stop = 12

    for i in range(1, my_num_rows+1):

        print_df = tmp_fa_df.iloc[start:stop].copy()

        print_df.index = range(1, print_df.shape[0]+1)

        # df.loc[df.index[#], 'NAME']

        print_df.loc[print_df.index[10], 'name'] = 'LibStd'

        print_df.loc[print_df.index[11], 'name'] = 'ladder'

        print_df.to_csv(ATTEMPT_DIR /
            f'FA_upload_{fabc}_row{row_list[i-1]}.csv', index=True, header=False)
        
        # print_df.to_csv(
        #     f'FA_upload_{fabc}_row{row_list[i-1]}.csv', index=True, header=False)

        start = start + 12

        stop = stop + 12


##########################
##########################

# ##########################
# ##########################
# def makeBioanalyzerInputFile(my_pippin_df):
#     # add info to start of barcode print file indicating the template and printer to use
#     x = '"Sample Name","Sample Comment","Rest. Digest","Observation"\n'

#     z = '"Chip Lot #","Reagent Kit Lot #"\n,\n\n"QC1 Min [%]","QC1 Max [%]","QC2 Min [%]","QC2 Max [%]"\n,,,\n\n"Chip Comment"\n\n\n"Study Name","Experimenter","Laboratory","Company","Department"\n,,,,\n\n"Study Comment"\n'

#     # y = '35 Curio,,,\nLadder,,,\n\n'

#     pool_list = my_pippin_df['Pool_Name'].tolist()

#     bc_file = open("Bioanalyzer_Input_round_A.txt", "w")

#     bc_file.writelines(x)

#     # bc_file.writelines(y)

#     for y in range(11):
#         if y < len(pool_list):
#             bc_file.writelines(f'{pool_list[y]},,,\n')
#         else:
#             bc_file.writelines(f'sample {y+1},,,\n')

#     bc_file.writelines('Ladder,,,\n\n')

#     bc_file.writelines(z)

#     return
# ##########################
# ##########################

##########################
# MAIN PROGRAM
##########################

PROJECT_DIR = Path.cwd()

ARCHIV_DIR = PROJECT_DIR / "archived_files"

POOL_DIR = PROJECT_DIR / "5_pooling"

FINISH_DIR = POOL_DIR / "D_finish_pooling"

REWORK_DIR = POOL_DIR / "E_pooling_and_rework"

ATTEMPT_DIR = REWORK_DIR / "Attempt_1"
ATTEMPT_DIR.mkdir(parents=True, exist_ok=True)

# get current working directory, which should be the project directory
prjct_dir = os.getcwd()

dir_name = "5_pooling"

prnt_dir = os.path.join(prjct_dir, dir_name)

sub_dir_name = "D_finish_pooling"

crnt_dir = os.path.join(prnt_dir, sub_dir_name)


# # pool creation .xls file is first argument when calling script
# pool_file = sys.argv[1]

pool_file = findPoolFile(crnt_dir)

############ THIS MUST BE RE-ACTIVATED ONCE FINISHED WITH TROUBLESHOOTING
# fix excel formulas by opening, saving, and closing file
# without changing anything
fixExcelFile(pool_file)


# create df from lib_info_submitted_to_clarity.db sqliute file
lib_df = readSQLdb(prjct_dir)

# add clarity pool ID to lib_info_submitted df
lib_df = addPoolClarityID(lib_df, pool_file)

# add pippin lane# and cassette barcode ot lib_df
lib_df = addPippinInfo(lib_df)



############ THIS MUST BE RE-ACTIVATED ONCE FINISHED WITH TROUBLESHOOTING
# fill out pool creation .xls sheet so that it can be uploaded to clarity
fillPoolCreationSheet(lib_df, pool_file)

# make transfer files for Hamilton for making pools
makePoolTransferFiles(lib_df)

# make tube barcode labels and return df that can be used for makeing pippin transfer files
pippin_df = makeTubeBarcodeFiles(lib_df)

# make pippin transfer files and add FA well positions to pippin_df
pippin_df, dest_list = makePippinTransferFiles(pippin_df)

# make pippin cassette barcode label
makePippinBarcodeFile(pippin_df)

# make input files for Fragment Analysis for reworked libs
makeFAinputFiles(pippin_df, dest_list)


# make temporary review file with lib_info_submitted with updated info about pooling info
lib_df.to_csv(FINISH_DIR / 'review_pool.csv', index=False, header=True)

# # add empty columns that will eventually hold results of FA QC
# pippin_df[['Passed_Pool','New_pool']] = ''

# # make new summary file for pool processing and rework
# pippin_df[['Pool_Name', 'Pool_Barcode', 'Pippin_Cassette', '1st_Pippin_lane','2nd_Pippin_lane','Dest_Tube_Size_Selected','FA_plate_barcode','FA_well','Passed_Pool','New_pool']].to_csv(POOLING_DIR /
#                                                                                   'pool_summary.csv', index=False, header=True)

# make new summary file for pool processing and rework
pippin_df[['Pool_Name', 'Pool_Barcode', 'Pippin_Cassette', '1st_Pippin_lane','2nd_Pippin_lane','Dest_Tube_Size_Selected','FA_plate_barcode','FA_well']].to_csv(POOL_DIR /
                                                                                  'pool_summary.csv', index=False, header=True)



# # create an input file for bioanalzyer run of size selected pools
# makeBioanalyzerInputFile(pippin_df)

# # get current date and time, will add to archive database file name
# date = datetime.now().strftime("%Y_%m_%d-Time%H-%M-%S")


# Path(PROJECT_DIR /
#      "lib_info_submitted_to_clarity.csv").rename(ARCHIV_DIR / f"archive_lib_info_submitted_to_clarity{date}.csv")
# # Path(ARCHIV_DIR / f"archive_lib_info_submitted_to_clarity_{date}.csv").touch()

# # create updated library info file
# lib_df.to_csv(PROJECT_DIR / 'lib_info_submitted_to_clarity.csv', index=False)

# Create success marker file to indicate script completed successfully
import os
os.makedirs('.workflow_status', exist_ok=True)
with open('.workflow_status/finish.pooling.libs.success', 'w') as f:
    f.write('Script completed successfully')
